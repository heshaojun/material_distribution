import openpyxl
import math


def load_data(sheet, mapper):
    """
    从excel中读取原始数据，
    :param sheet: excel 工作表对象
    :param mapper : 数据映射
    :return: {"仓库编号":{"物料编号":[但前库存,月需求量,库存天数,需求量]}}
    """
    data_dict = {}
    row_index = 3
    while True:
        # 读取仓库编号
        warehouse_no = sheet.cell(row=row_index, column=1).value
        # 读取物料编号
        material_no = sheet.cell(row=row_index, column=2).value
        if not warehouse_no or not material_no:
            break
        if not mapper.__contains__(material_no):
            continue
        msg = mapper[material_no]
        if not data_dict.__contains__(warehouse_no):
            data_dict[warehouse_no] = {str(material_no): []}
        # 获取当前库存
        stock = int(sheet.cell(row=row_index, column=4).value)
        # 获取月需求量
        forecast = int(sheet.cell(row=row_index, column=5).value)
        # 请求库存天数
        dos = int(math.ceil(float(stock) * 30 / float(forecast)))  # 向上小数去尾
        demand = 0  # 需求量
        warehouse_ = data_dict[warehouse_no]
        if msg[1] > dos:
            demand = forecast * msg[2] - stock
        warehouse_[material_no] = [stock, forecast, dos, demand]
        row_index += 1
    return data_dict


def record_data(sheet, result_data):
    """
    纪录处理结果数据
    :param sheet: 需要纪录数据当新工作表
    :param result_data: 结果数据
    :return:
    """
    pass


def compute_demand(raw_data):
    """
    计算单一物料最大月需求量
    :param raw_data: 原始数据{"仓库编号":{"物料编号":[但前库存,月需求量,库存天数,需求量]}}
    :return: 返回全部物料的最需求量
    """
    material_group = {}
    for k_, m_ in raw_data.items():
        for n_, d_ in m_.items():
            if material_group.__contains__(n_):
                material_group[n_] = material_group.get(n_) + d_[3]
            else:
                material_group[n_] = d_[3]

    return {k: math.floor(v) for k, v in material_group.items()}  # 向下取整


def compute_average_dos(raw_data, assignable):
    """
    计算总共物料可供给天数均值
    :param raw_data: 原始数据
    :param assignable: 可分配物料数量
    :return: 物料分配后可可持续天数的均值
    """

    stock_count = {}  # 存储每种物料的总库存
    forecast_count = {}  # 存储每种物料的月需
    for h_, m_ in raw_data.items():
        for n_, d_ in m_.items():
            if stock_count.__contains__(n_):
                stock_count[n_] = stock_count[n_] + d_[0]
            else:
                stock_count[n_] = d_[0]
            if forecast_count.__contains__(n_):
                forecast_count[n_] = forecast_count[n_] + d_[1]
            else:
                forecast_count[n_] = d_[1]
    return {k: (stock_count[k] + assignable[k]) * 30 / v for k, v in forecast_count}


def compute_assignable(allocation, max_demand):
    """
    计算每种物料预计可分配数量
    :param allocation:
    :param max_demand:
    :return:
    """
    assignable = {}
    # 获取每种物料的预分配数量
    for m_, d_ in max_demand.items():
        a_ = allocation.get(m_)
        if a_ < d_:
            a_ = d_
        assignable[m_] = a_
    return assignable


if __name__ == "__main__":
    # 物料字典，{"物料编号":[物料可分配量,当前物料可分配上限天数,物料最大供给月数]}
    material_dic = {"_1": [8, 45, 2], "_2": [20, 45, 1.5], "_3": [8, 45, 2]}
    # 加载excel文件
    wb = openpyxl.load_workbook("material.xlsx")
    input_sheet = wb['input']
    raw_data_ = load_data(input_sheet, material_dic)
    max_material_demand_ = compute_demand(raw_data_)
    print(raw_data_)
    print(max_material_demand_)
    print("hello world!")
    print(math.ceil(2))
